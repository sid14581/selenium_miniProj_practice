import openpyxl


class HomePageData:
    @staticmethod
    def get_testData(test_case_name):
        exe_dict = {}

        book = openpyxl.load_workbook("C:\\Users\\siddh\\OneDrive\\Desktop\\demoData.xlsx")
        sheet = book.active

        # cell = sheet.cell(row=1, column=2)
        # sheet.cell(row=2, column=2).value = "Sid"
        # sheet['C3'].value = "harsh@gmail.com"
        # book.save("C:\\Users\\siddh\\OneDrive\\Desktop\\demoData.xlsx")

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    exe_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [exe_dict]

# print(HomePageData.get_testData())
